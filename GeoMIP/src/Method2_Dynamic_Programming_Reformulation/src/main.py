# from src.controllers.manager import Manager

# from src.controllers.strategies.force import BruteForce
# from src.controllers.strategies.q_nodes import QNodes
# from src.controllers.strategies.geometric import GeometricSIA


# def iniciar():
#     """Punto de entrada principal"""
#                     # ABCD #
#     # estado_inicial = "100"
#     # condiciones =    "111"
#     # alcance =        "111"
#     # mecanismo =      "111"
#     # estado_inicial = "0000"
#     # condiciones =    "1111"
#     # alcance =        "1111"
#     # mecanismo =      "1111"
#     # estado_inicial = "1000"
#     # condiciones =    "1111"
#     # alcance =        "0111"
#     # mecanismo =      "1111"
#     # estado_inicial = "100000"
#     # condiciones =    "111111"
#     # alcance =        "101011"
#     # mecanismo =      "111111"
#     # estado_inicial = "100000"
#     # condiciones =    "111111"
#     # alcance =        "111111"
#     # mecanismo =      "111111"
#     # estado_inicial = "100000"
#     # condiciones =    "111111"
#     # alcance =        "111111"
#     # mecanismo =      "011111"
#     # estado_inicial = "1000000000"
#     # condiciones =    "1111111111"
#     # alcance =        "1111111111"
#     # mecanismo =      "1111111111"
#     estado_inicial = "1000000000"
#     condiciones =    "1111111111"
#     alcance =        "0101010101"
#     mecanismo =      "1111111111"
#     # estado_inicial = "1000000000"
#     # condiciones =    "1111111111"
#     # alcance =        "1111111110"
#     # mecanismo =      "1111111111"
#     # estado_inicial = "10000000000000000000"
#     # condiciones =    "11111111111111111111"
#     # alcance =        "11111111111111111111"
#     # mecanismo =      "11111111111111111111"
#     # estado_inicial = "10000000000000000000"
#     # condiciones =    "11111111111111111111"
#     # alcance =        "11011011011011011011"
#     # mecanismo =      "10101010101010101010"

#     gestor_sistema = Manager(estado_inicial)

#     ### Ejemplo de solución mediante módulo de fuerza bruta ###
#     analizador_fb = GeometricSIA(gestor_sistema)
#     # analizador_fb = BruteForce(gestor_sistema)
#     sia_uno = analizador_fb.aplicar_estrategia(
#         condiciones,
#         alcance,
#         mecanismo,
#     )
#     print(sia_uno)
import traceback

from openpyxl import load_workbook

from src.controllers.strategies.k_geometric import KGeoMIPStrategy
from src.controllers.manager import Manager
from src.controllers.strategies.geometric import GeometricSIA
# Optional import: this project often runs only geometric strategy.
try:
    from src.controllers.strategies.phi import Phi
except Exception:
    Phi = None
import multiprocessing
import numpy as np
import pandas as pd
import os
import re
from pathlib import Path


METHOD2_ROOT = Path(__file__).resolve().parents[1]
GEOMIP_ROOT = Path(__file__).resolve().parents[3]

def convertir_a_binario(texto, n_bits=20):
    posiciones = "ABCDEFGHIJKLMNOPQRST"[:n_bits]
    binario = ["0"] * n_bits
    for letra in texto:
        if letra in posiciones:
            binario[posiciones.index(letra)] = "1"
    return "".join(binario)

def ejecutar_con_tiempo(config_sistema, condiciones, alcance, mecanismo, resultado_queue, tpm, k):
    print("ENTRE A EJECUTAR_CON_TIEMPO")
    print("alcance =", alcance)
    print("mecanismo =", mecanismo)
    try:

        analizador_fi = KGeoMIPStrategy(config_sistema)

        resultado = analizador_fi.aplicar_estrategia(
            condiciones,
            alcance,
            mecanismo,
            tpm,
            k=k
        )

        resultado_queue.put({
            "particion": resultado.particion,
            "perdida": str(resultado.perdida).replace(".", ","),
            "tiempo": str(resultado.tiempo_ejecucion).replace(".", ","),
        })

    except Exception as e:

        print("ERROR:")
        traceback.print_exc()

        resultado_queue.put({
            "particion": None,
            "perdida": None,
            "tiempo": None,
        })
            
def resolver_tpm_path(estado_inicio: str) -> Path:
    """Find TPM file in common project locations based on state size."""
    sample_name = f"N{len(estado_inicio)}A.csv"
    candidates = (
        METHOD2_ROOT / "src" / ".samples" / sample_name,
        METHOD2_ROOT / ".samples" / sample_name,
        GEOMIP_ROOT / "data" / "samples" / sample_name,
    )
    for candidate in candidates:
        if candidate.exists():
            return candidate
    raise FileNotFoundError(
        f"No se encontró la TPM '{sample_name}'. Busqué en: {', '.join(str(c) for c in candidates)}"
    )


def inferir_estado_inicial(tamaño) -> str:
    """Infer an initial state from available datasets (prefers largest NxA.csv)."""
    sample_dirs = (
        METHOD2_ROOT / "src" / ".samples",
        METHOD2_ROOT / ".samples",
        GEOMIP_ROOT / "data" / "samples",
    )
    pattern = re.compile(r"N(\d+)[A-Z]\.csv$")

    
    return "1" + ("0" * (tamaño - 1))


def ejecutar_desde_excel(
    ruta_excel: Path,
    hoja=4,
    k=2,
    inicio=0,
    cantidad=50,
    tamaño_estado=22,
    estado_inicio: str | None = None,
    condiciones: str | None = None,
):

    print(f"Procesando archivo Excel: {ruta_excel}")

    wb = load_workbook(ruta_excel)

    ws = wb.worksheets[hoja] if isinstance(hoja, int) else wb[hoja]

    print(f"Procesando hoja: {ws.title}")

    COLUMNAS = {
        2: {"particion": "G", "perdida": "H", "tiempo": "I"},
        3: {"particion": "M", "perdida": "N", "tiempo": "O"},
        4: {"particion": "S", "perdida": "T", "tiempo": "U"},
        5: {"particion": "Y", "perdida": "Z", "tiempo": "AA"},

    }

    fila_inicio = 6 + inicio
    fila_fin = min(ws.max_row, fila_inicio + cantidad - 1)

    estado_inicio = estado_inicio or inferir_estado_inicial(tamaño_estado)

    print(f"Estado inicial inferido: {estado_inicio}")

    condiciones = condiciones or ("1" * len(estado_inicio))

    tpm_path = resolver_tpm_path(estado_inicio)
    print(f"TPM cargada desde {tpm_path}")
    tpm = np.genfromtxt(tpm_path, delimiter=",")
    cols = COLUMNAS[k]
    print(f"Procesando tmp cargada")

    for fila_excel in range(fila_inicio, fila_fin + 1):

        alcance = ws[f"B{fila_excel}"].value
        mecanismo = ws[f"C{fila_excel}"].value
        if alcance is None or mecanismo is None:
            continue
        print(f"Alcance: {alcance}, Mecanismo: {mecanismo}")
        
        alcance = convertir_a_binario(alcance, n_bits=len(estado_inicio))
        mecanismo = convertir_a_binario(mecanismo, n_bits=len(estado_inicio))
        print(f"Alcance: {alcance}, Mecanismo: {mecanismo}")
       
        print(
            f"Fila {fila_excel} "
            f"Alcance={alcance} "
            f"Mecanismo={mecanismo}"
        )

        config_sistema = Manager(estado_inicial=estado_inicio)

        resultado_queue = multiprocessing.Queue()

        proceso = multiprocessing.Process(
            target=ejecutar_con_tiempo,
            args=(
                config_sistema,
                condiciones,
                alcance,
                mecanismo,
                resultado_queue,
                tpm,
                k
            )
        )

        proceso.start()

        proceso.join(timeout=3600)

        if proceso.is_alive():

            print(f"Fila {fila_excel} tiempo límite alcanzado")

            proceso.terminate()
            proceso.join()

            resultado = {
                "particion": None,
                "perdida": None,
                "tiempo": None
            }

        else:

            resultado = (
                resultado_queue.get()
                if not resultado_queue.empty()
                else {
                    "particion": None,
                    "perdida": None,
                    "tiempo": None
                }
            )
        print(
            f"Resultado fila {fila_excel}: "
            f"Partición={resultado['particion']} "
            f"Pérdida={resultado['perdida']} "
            f"Tiempo={resultado['tiempo']}"
        )
        ws[f"{cols['particion']}{fila_excel}"] = resultado["particion"]
        ws[f"{cols['perdida']}{fila_excel}"] = resultado["perdida"]
        ws[f"{cols['tiempo']}{fila_excel}"] = resultado["tiempo"]

        wb.save(ruta_excel)

    wb.save(ruta_excel)

    print(f"Resultados guardados en {ruta_excel}")
def iniciar():
    ruta_excel = Path(
    os.getenv(
        "GEOMIP_INPUT_XLSX",
        str(GEOMIP_ROOT / "results" / "DatosPruebas2026_1.xlsx"),
        )
    )
    for k in range(2, 6):
        ejecutar_desde_excel(
            ruta_excel=ruta_excel,
            hoja=5,
            k=k,
            tamaño_estado=25,
            cantidad=50
        )


