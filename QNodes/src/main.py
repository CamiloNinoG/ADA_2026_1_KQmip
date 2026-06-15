import os
import time
import pandas as pd
from datetime import datetime
import openpyxl
from src.controllers.manager import Manager
from src.strategies.kqnodes.kqnodes import KQNodes
from src.strategies.q_nodes import QNodes
from src.strategies.force import BruteForce
from src.strategies.kqnodes.profilter import KQNodesProfiler

# =========================================================================
# CONFIGURACIÓN DINÁMICA
# =========================================================================
LONGITUD_ELEMENTOS = 10
nombre_hoja = f"{LONGITUD_ELEMENTOS}A-Elementos"
# Método a usar
estado_inicial = "1" + "0" * (LONGITUD_ELEMENTOS - 1)  # Ej: "1000000000" para 10b
condiciones = "1" * LONGITUD_ELEMENTOS  # Ej: "1111111111" para 10b
gestor_redes = Manager(estado_inicial)
mpt = gestor_redes.cargar_red()
analizador_bf = KQNodes(mpt)  # "KQNodes", "QNodes", "BruteForce"


# =========================================================================
def obtener_timestamp():
    """Devuelve la hora actual formateada para los prints."""
    return datetime.now().strftime("%H:%M:%S.%f")[:-3]


def texto_a_binario(texto, universo):
    """Convierte una cadena de letras a su binario adaptado al tamaño del universo."""
    if pd.isna(texto):
        return "0" * len(universo)
    resultado = "".join(
        ["1" if letra in str(texto).upper() else "0" for letra in universo]
    )
    return resultado


def iniciar():
    print(
        f"[{obtener_timestamp()}] 🚀 Iniciando proceso de automatización estructurado..."
    )
    universo_letras = "".join([chr(65 + i) for i in range(LONGITUD_ELEMENTOS)])
    ruta_excel = os.path.join("src", "results", "pruebas.xlsx")
    ruta_salida = os.path.join("src", "results", "pruebas_con_resultados.xlsx")
    if not os.path.exists(ruta_excel):
        print(
            f"[{obtener_timestamp()}] ❌ ERROR: No se encontró el archivo Excel en '{ruta_excel}'"
        )
        return
    print(
        f"[{obtener_timestamp()}] 📊 Cargando datos con Pandas para el procesamiento lógico..."
    )
    df_origen = pd.read_excel(ruta_excel, sheet_name=nombre_hoja, skiprows=4)
    df_origen.columns = df_origen.columns.str.strip().str.replace("\n", " ")
    print(
        f"[{obtener_timestamp()}] 📊 Total de filas detectadas por Pandas: {len(df_origen)}"
    )
    print(f"[{obtener_timestamp()}] 📋 Abriendo plantilla con openpyxl...")
    wb = openpyxl.load_workbook(ruta_excel)
    if nombre_hoja not in wb.sheetnames:
        print(f"[{obtener_timestamp()}] ❌ ERROR: La hoja '{nombre_hoja}' no existe.")
        return
    sheet = wb[nombre_hoja]
    print(
        f"[{obtener_timestamp()}] 🌐 Inicializando Gestor de Redes (Manager) a {LONGITUD_ELEMENTOS} bits..."
    )
    print(f"[{obtener_timestamp()}] 🧠 Inicializando Analizador de (KQNodes)...")
    KQNodesProfiler.wrap(analizador_bf)
    print(f"\n[{obtener_timestamp()}] 🔍 Procesando e inyectando resultados...")
    print("=" * 80)
    pruebas_procesadas = 0
    # Iterar por cada fila del DataFrame original
    for index, fila in df_origen.iterrows():
        id_prueba = fila.get("#Prueba")
        alcance_txt = fila.get("Alcance o Purview (t+1)")
        mecanismo_txt = fila.get("Mecanismo(t)")
        if pd.isna(alcance_txt) or pd.isna(mecanismo_txt):
            print(
                f"[{obtener_timestamp()}] ⏭️  Fila {index} saltada porque Alcance o Mecanismo están vacíos."
            )
            continue
        pruebas_procesadas += 1
        num_prueba = int(id_prueba) if not pd.isna(id_prueba) else (index + 1)
        fila_excel = index + 6
        alcance_bin = texto_a_binario(alcance_txt, universo_letras)
        mecanismo_bin = texto_a_binario(mecanismo_txt, universo_letras)
        print(
            f"[{obtener_timestamp()}] 🕒 [Prueba #{num_prueba}] (Fila Excel: {fila_excel})"
        )
        print(
            f"      📥 Datos -> Alcance: {alcance_txt} ({alcance_bin}) | Mecanismo: {mecanismo_txt} ({mecanismo_bin})"
        )
        # 🔄 LOOP SOLICITADO: De K = 3 hasta 4 (K=3) para la misma fila
        for k in range(2, 6):
            print(f"      ⚡ Ejecutando 'aplicar_estrategia_k' con K = {k}...")
            tiempo_inicio = time.time()
            try:
                analizador_bf.reset_estado()
                sia_cero = analizador_bf.aplicar_estrategia_k(
                    estado_inicial,
                    condiciones,
                    alcance_bin,
                    mecanismo_bin,
                    k,  # Inyección dinámica de K
                )
                tiempo_fin = time.time() - tiempo_inicio
                particion_val = getattr(sia_cero, "particion", "N/A")
                perdida_val = getattr(
                    sia_cero, "phi", getattr(sia_cero, "perdida", sia_cero)
                )
                tiempo_str = f"{tiempo_fin:.4f}s"
                print(
                    f"      ✨ [OK K={k}] {tiempo_str} | Partición: {particion_val} | Pérdida: {perdida_val}"
                )
            except Exception as e:
                tiempo_fin = time.time() - tiempo_inicio
                particion_val = "ERROR"
                perdida_val = str(e)
                tiempo_str = f"{tiempo_fin:.4f}s"
                print(f"      ❌ ERROR en ejecución K={k}: {e}")
            # 🗺️ Mapeo e Inyección dinámica de columnas en base a K:
            # K=2 -> Col D (4), E (5), F (6)
            # K=3 -> Col G (7), H (8), I (9)
            # K=4 -> Col J (10), K (11), L (12)
            # K=5 -> Col M (13), N (14), O (15)
            col_base = 4 + (k - 2) * 3
            sheet.cell(row=fila_excel, column=col_base, value=str(particion_val))

            # Guardamos la pérdida como tipo float nativo para mantener precisión ajustable en Excel
            try:
                sheet.cell(
                    row=fila_excel, column=col_base + 1, value=float(perdida_val)
                )
            except (ValueError, TypeError):
                sheet.cell(row=fila_excel, column=col_base + 1, value=str(perdida_val))
            sheet.cell(row=fila_excel, column=col_base + 2, value=tiempo_str)
    print("\n" + "=" * 80)
    print(
        f"[{obtener_timestamp()}] 🏁 Procesamiento completo. Total procesadas con éxito: {pruebas_procesadas}"
    )
    if pruebas_procesadas > 0:
        print(
            f"[{obtener_timestamp()}] 💾 Cambios aplicados en memoria. Guardando archivo final..."
        )
        wb.save(ruta_salida)
    else:
        print(
            f"[{obtener_timestamp()}] ⚠️ No se procesaron filas legítimas. El archivo de salida no fue modificado."
        )
    wb.close()
    print(f"[{obtener_timestamp()}] 📁 Proceso terminado.")
    KQNodesProfiler.report()


if __name__ == "__main__":
    iniciar()
